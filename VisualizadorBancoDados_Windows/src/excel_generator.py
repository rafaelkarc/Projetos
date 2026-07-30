import pandas as pd
import openpyxl
from openpyxl import load_workbook
from calendar import monthrange
import os
from datetime import datetime
import math

def update_existing_spreadsheet(file_path, sheet_name, cell, value):
    """
    Atualiza uma célula específica em uma planilha Excel existente.
    
    Args:
        file_path (str): Caminho para o arquivo Excel
        sheet_name (str): Nome da aba/sheet
        cell (str): Célula de destino (ex: 'A1', 'B2')
        value: Valor a ser inserido na célula
    
    Returns:
        bool: True se a operação foi bem-sucedida, False caso contrário
    """
    try:
        # Verificar se o arquivo existe
        if not os.path.exists(file_path):
            print(f"Arquivo não encontrado: {file_path}. Tentando usar {PLANILHA_PU_FILE}")
            file_path = PLANILHA_PU_FILE
            if not os.path.exists(file_path):
                print(f"Arquivo {PLANILHA_PU_FILE} também não encontrado. Não é possível salvar.")
                return False
        
        # Carregar a planilha
        workbook = load_workbook(file_path)
        
        # Verificar se a aba existe
        if sheet_name not in workbook.sheetnames:
            print(f"Aba '{sheet_name}' não encontrada. Abas disponíveis: {workbook.sheetnames}")
            return False
        
        # Selecionar a aba
        worksheet = workbook[sheet_name]
        
        # Atualizar a célula
        worksheet[cell] = value
        
        # Salvar o arquivo
        workbook.save(file_path)
        workbook.close()
        
        print(f"Valor '{value}' salvo com sucesso na célula {cell} da aba '{sheet_name}'")
        return True
        
    except Exception as e:
        print(f"Erro ao atualizar planilha: {str(e)}")
        return False

def save_calculation_to_spreadsheet(file_path, vna_calculado, fator_correcao, data_calculo=None):
    """
    Salva os resultados do cálculo VNA na planilha em células específicas.
    
    Args:
        file_path (str): Caminho para o arquivo Excel
        vna_calculado (float): Valor do VNA calculado
        fator_correcao (float): Fator de correção
        data_calculo (str): Data do cálculo (opcional)
    
    Returns:
        bool: True se a operação foi bem-sucedida, False caso contrário
    """
    try:
        # Definir as células onde os dados serão salvos
        # Você pode ajustar essas células conforme necessário
        celula_vna = 'M3'  # Célula para VNA calculado
        celula_fator = 'M4'  # Célula para fator de correção
        celula_data = 'M5'  # Célula para data do cálculo
        
        # Nome da aba onde salvar (ajuste conforme necessário)
        sheet_name = 'PU'
        
        # Salvar VNA calculado
        success_vna = update_existing_spreadsheet(file_path, sheet_name, celula_vna, vna_calculado)
        
        # Salvar fator de correção
        success_fator = update_existing_spreadsheet(file_path, sheet_name, celula_fator, fator_correcao)
        
        # Salvar data do cálculo se fornecida
        success_data = True
        if data_calculo:
            success_data = update_existing_spreadsheet(file_path, sheet_name, celula_data, data_calculo)
        
        return success_vna and success_fator and success_data
        
    except Exception as e:
        print(f"Erro ao salvar cálculo na planilha: {str(e)}")
        return False

def save_pu_par_to_spreadsheet(file_path, vna, taxa_juros, dias_uteis, base_calculo, pu_par, data_calculo=None):
    """
    Salva os resultados do cálculo PU PAR na planilha.
    
    Args:
        file_path (str): Caminho para o arquivo Excel
        vna (float): Valor Nominal Atualizado
        taxa_juros (float): Taxa de Juros (% a.a.)
        dias_uteis (int): Dias Úteis desde o Último Pagamento
        base_calculo (int): Base de Cálculo (252/360/365 dias)
        pu_par (float): PU PAR calculado
        data_calculo (str): Data do cálculo (opcional)
    
    Returns:
        bool: True se a operação foi bem-sucedida, False caso contrário
    """
    try:
        # sheet_name = 'PU PAR' # Antigo nome da aba
        sheet_name = '22F1187231' # Novo nome da aba
        if not os.path.exists(file_path):
            print(f"Arquivo '{file_path}' não encontrado. Usando o arquivo padrão '{PLANILHA_PU_FILE}'.")
            file_path = PLANILHA_PU_FILE
        
        # Carregar a planilha
        workbook = load_workbook(file_path)
        
        # Verificar se a aba existe, se não, criar
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
            workbook.save(file_path)
            workbook.close()
            # Recarregar o workbook após criar a sheet
            workbook = load_workbook(file_path)

        worksheet = workbook[sheet_name]

        # Encontrar a linha da célula 'RESIDUAL' na coluna N
        residual_row = -1
        for row_idx in range(1, worksheet.max_row + 1):
            cell_value = worksheet[f'N{row_idx}'].value
            if isinstance(cell_value, str) and cell_value.strip().upper() == 'RESIDUAL':
                residual_row = row_idx
                break
        
        if residual_row == -1:
            print(f"Célula 'RESIDUAL' não encontrada na coluna N da aba '{sheet_name}'. Salvando na primeira célula disponível na coluna N.")
            # Se 'RESIDUAL' não for encontrado, salva na próxima linha vazia da coluna N
            next_row = worksheet.max_row + 1
        else:
            next_row = residual_row + 1

        # Salvar o PU PAR na célula abaixo de 'RESIDUAL' na coluna N
        success_pu_par = update_existing_spreadsheet(file_path, sheet_name, f'N{next_row}', pu_par)

        # Remover o salvamento dos outros dados do PU PAR na aba 'PU PAR' se ela ainda existir
        # Estes dados não serão mais salvos na aba 'PU PAR' para evitar duplicação ou confusão
        # if 'PU PAR' in workbook.sheetnames:
        #     # Pode-se optar por limpar a aba 'PU PAR' ou simplesmente não escrever nela
        #     pass

        # Os dados de entrada (VNA, Taxa de Juros, Dias Úteis, Base de Cálculo) não serão salvos nesta aba específica
        # Se for necessário salvá-los, deve-se criar uma lógica separada para isso em outra aba ou local.
        
        return success_pu_par
        
    except Exception as e:
        print(f"Erro ao salvar PU PAR na planilha: {str(e)}")
        return False

def save_pu_operacao_to_spreadsheet(file_path, taxa_mercado, base_calculo, fluxos, pu_operacao, data_calculo=None):
    """
    Salva o PU Operação na aba '22F1187231' da planilha_pu.xlsx.
    Primeiro tenta encontrar uma linha com a data do dia 25 do mês atual.
    Se não encontrar, adiciona uma nova linha no final da planilha.
    """
    from openpyxl import load_workbook
    from datetime import datetime

    try:
        sheet_name = '22F1187231'
        if not os.path.exists(file_path):
            print(f"Arquivo '{file_path}' não encontrado.")
            return False

        workbook = load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            print(f"Aba '{sheet_name}' não encontrada na planilha.")
            return False

        worksheet = workbook[sheet_name]

        # Data atual
        hoje = datetime.now()
        mes_atual = hoje.month
        ano_atual = hoje.year

        # Buscar na coluna D a linha onde a data seja 25 do mês atual e do ano atual
        linha_encontrada = None
        for row in range(1, worksheet.max_row + 1):
            cell_value = worksheet[f'D{row}'].value
            if isinstance(cell_value, datetime):
                if cell_value.day == 25 and cell_value.month == mes_atual and cell_value.year == ano_atual:
                    linha_encontrada = row
                    break

        if linha_encontrada is None:
            # Se não encontrou a data específica, adicionar uma nova linha
            print(f"Data 25/{mes_atual:02d}/{ano_atual} não encontrada. Adicionando nova linha.")
            
            # Encontrar a próxima linha vazia
            proxima_linha = worksheet.max_row + 1
            
            # Adicionar a data na coluna D
            data_para_salvar = datetime(ano_atual, mes_atual, 25)
            worksheet[f'D{proxima_linha}'] = data_para_salvar
            
            # Usar a nova linha
            linha_encontrada = proxima_linha

        # Escrever PU Operação na coluna N da linha encontrada
        worksheet[f'N{linha_encontrada}'] = pu_operacao
        
        # Adicionar informações adicionais se necessário
        if data_calculo:
            # Salvar data do cálculo na coluna O (opcional)
            worksheet[f'O{linha_encontrada}'] = data_calculo

        workbook.save(file_path)
        workbook.close()

        print(f"PU Operação {pu_operacao} salvo com sucesso na célula N{linha_encontrada}.")
        return True

    except Exception as e:
        print(f"Erro ao salvar PU Operação: {str(e)}")
        return False

def save_filtros_estatisticos_to_spreadsheet(file_path, taxas_originais, taxas_filtradas, data_calculo=None):
    """
    Salva os resultados dos filtros estatísticos na planilha.
    
    Args:
        file_path (str): Caminho para o arquivo Excel
        taxas_originais (list): Lista de taxas originais
        taxas_filtradas (list): Lista de taxas após filtros
        data_calculo (str): Data do cálculo (opcional)
    
    Returns:
        bool: True se a operação foi bem-sucedida, False caso contrário
    """
    try:
        sheet_name = 'Filtros Estatisticos'
        
        # Criar a aba se não existir
        workbook = load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
            workbook.save(file_path)
            workbook.close()
        
        # Salvar taxas originais
        success_originais = True
        for i, taxa in enumerate(taxas_originais):
            row = 2 + i
            success_originais &= update_existing_spreadsheet(file_path, sheet_name, f'A{row}', taxa)
        
        # Salvar taxas filtradas
        success_filtradas = True
        for i, taxa in enumerate(taxas_filtradas):
            row = 2 + i
            success_filtradas &= update_existing_spreadsheet(file_path, sheet_name, f'B{row}', taxa)
        
        success_data = True
        if data_calculo:
            success_data = update_existing_spreadsheet(file_path, sheet_name, 'C1', data_calculo)
        
        return success_originais and success_filtradas and success_data
        
    except Exception as e:
        print(f"Erro ao salvar Filtros Estatísticos na planilha: {str(e)}")
        return False

def calcular_pu_par(vna, taxa_juros, dias_uteis, base_calculo):
    """
    Calcula o PU PAR baseado nos parâmetros fornecidos.
    
    Args:
        vna (float): Valor Nominal Atualizado
        taxa_juros (float): Taxa de Juros (% a.a.)
        dias_uteis (int): Dias Úteis desde o Último Pagamento
        base_calculo (int): Base de Cálculo (252/360/365 dias)
    
    Returns:
        float: PU PAR calculado
    """
    try:
        # Converter taxa anual para decimal
        taxa_decimal = taxa_juros / 100
        
        # Calcular fator de juros
        fator_juros = (1 + taxa_decimal) ** (dias_uteis / base_calculo)
        
        # Calcular PU PAR
        pu_par = vna * fator_juros
        
        return round(pu_par, 8)
        
    except Exception as e:
        print(f"Erro ao calcular PU PAR: {str(e)}")
        return None

def calcular_pu_operacao(taxa_mercado, base_calculo, fluxos):
    """
    Calcula o PU Operação baseado na taxa de mercado e fluxos de pagamento.
    
    Args:
        taxa_mercado (float): Taxa de Mercado (% a.a.)
        base_calculo (int): Base de Cálculo
        fluxos (list): Lista de fluxos de pagamento
    
    Returns:
        float: PU Operação calculado
    """
    try:
        # Converter taxa anual para decimal
        taxa_decimal = taxa_mercado / 100
        
        pu_operacao = 0
        
        for fluxo in fluxos:
            dias_uteis = fluxo.get('dias_uteis', 0)
            valor = fluxo.get('valor', 0)
            
            # Calcular valor presente do fluxo
            fator_desconto = (1 + taxa_decimal) ** (dias_uteis / base_calculo)
            valor_presente = valor / fator_desconto
            
            pu_operacao += valor_presente
        
        return round(pu_operacao, 8)
        
    except Exception as e:
        print(f"Erro ao calcular PU Operação: {str(e)}")
        return None

def aplicar_filtros_estatisticos(taxas):
    """
    Aplica filtros estatísticos para remover outliers das taxas.
    
    Args:
        taxas (list): Lista de taxas
    
    Returns:
        list: Lista de taxas filtradas
    """
    try:
        if not taxas or len(taxas) < 3:
            return taxas
        
        # Converter para números
        taxas_numericas = [float(taxa) for taxa in taxas if taxa != '']
        
        if len(taxas_numericas) < 3:
            return taxas_numericas
        
        # Calcular estatísticas
        media = sum(taxas_numericas) / len(taxas_numericas)
        variancia = sum((x - media) ** 2 for x in taxas_numericas) / len(taxas_numericas)
        desvio_padrao = math.sqrt(variancia)
        
        # Filtrar outliers (valores fora de 2 desvios padrão)
        limite_inferior = media - 2 * desvio_padrao
        limite_superior = media + 2 * desvio_padrao
        
        taxas_filtradas = [taxa for taxa in taxas_numericas 
                          if limite_inferior <= taxa <= limite_superior]
        
        return taxas_filtradas
        
    except Exception as e:
        print(f"Erro ao aplicar filtros estatísticos: {str(e)}")
        return taxas

def convert_xls_to_xlsx(xls_file_path):
    """
    Converte um arquivo .xls para .xlsx para melhor compatibilidade.
    
    Args:
        xls_file_path (str): Caminho para o arquivo .xls
    
    Returns:
        str: Caminho para o arquivo .xlsx convertido
    """
    try:
        # Ler o arquivo .xls
        df_dict = pd.read_excel(xls_file_path, sheet_name=None)
        
        # Criar o nome do arquivo .xlsx
        xlsx_file_path = xls_file_path.replace('.xls', '.xlsx')
        
        # Salvar como .xlsx
        with pd.ExcelWriter(xlsx_file_path, engine='openpyxl') as writer:
            for sheet_name, df in df_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"Arquivo convertido de .xls para .xlsx: {xlsx_file_path}")
        return xlsx_file_path
        
    except Exception as e:
        print(f"Erro ao converter arquivo: {str(e)}")
        return None

# Variável global para o nome do arquivo da planilha
PLANILHA_PU_FILE = "planilha_pu.xlsx"

if __name__ == "__main__":
    # Teste da função
    file_path = "planilha_pu.xls"
    
    # Converter para .xlsx se necessário
    xlsx_path = convert_xls_to_xlsx(file_path)
    
    if xlsx_path:
        # Testar salvamento
        test_success = save_calculation_to_spreadsheet(
            xlsx_path, 
            1073.37, 
            1.07337126, 
            datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        )
        
        if test_success:
            print("Teste de salvamento bem-sucedido!")
        else:
            print("Falha no teste de salvamento.")





