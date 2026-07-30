"""
Módulo para cálculo de amortização
Implementa a fórmula: =TRUNCAR(G * PROCV(D; 'Amort TS'!G3:H140; 2; 0); 8)
"""

import openpyxl
from decimal import Decimal, ROUND_DOWN
from typing import Optional, Dict, Tuple
import math
import os


class AmortizacaoCalculator:
    """Calculadora de Amortização integrada com a planilha_pu.xlsx"""
    
    def __init__(self, planilha_path: str):
        """
        Inicializa a calculadora com o caminho da planilha
        
        Args:
            planilha_path: Caminho para o arquivo planilha_pu.xlsx
        """
        self.planilha_path = planilha_path
        self.wb = None
        self.load_workbook()
    
    def load_workbook(self):
        """Carrega a planilha"""
        try:
            # Carregar com data_only=True para obter valores calculados
            self.wb = openpyxl.load_workbook(self.planilha_path, data_only=True)
            print(f"✅ Planilha carregada: {self.planilha_path}")
        except Exception as e:
            print(f"❌ Erro ao carregar planilha: {e}")
            raise
    
    def procv_amort_ts(self, chave: str) -> Optional[float]:
        """
        Implementa PROCV na aba 'Amort TS'
        Busca o valor da chave na coluna G e retorna o valor correspondente da coluna H
        
        Equivalente a: =PROCV(chave; 'Amort TS'!G3:H140; 2; 0)
        
        Args:
            chave: Valor a buscar na coluna G (ex: data formatada)
            
        Returns:
            Valor correspondente da coluna H ou None se não encontrado
        """
        if 'Amort TS' not in self.wb.sheetnames:
            raise ValueError("Aba 'Amort TS' não encontrada na planilha")
        
        ws_amort = self.wb['Amort TS']
        
        # Procurar pela chave na coluna G (começando da linha 3)
        for row in range(3, 141):  # G3:H140
            cell_g = ws_amort[f'G{row}'].value
            
            # Comparar a chave (pode ser string ou data)
            if cell_g is not None:
                # Converter para string para comparação
                cell_g_str = str(cell_g).strip()
                chave_str = str(chave).strip()
                
                if cell_g_str == chave_str:
                    # Encontrou! Retornar valor da coluna H
                    cell_h = ws_amort[f'H{row}'].value
                    if cell_h is not None:
                        return float(cell_h)
                    else:
                        return None
        
        # Não encontrou
        return None
    
    def truncate_decimal(self, value: Decimal, decimals: int = 8) -> Decimal:
        """
        Trunca um Decimal para um número específico de casas decimais
        (sem arredondar)
        
        Args:
            value: Valor Decimal a truncar
            decimals: Número de casas decimais
            
        Returns:
            Valor truncado como Decimal
        """
        if value is None:
            return None
        
        # Converter para Decimal se necessário
        if not isinstance(value, Decimal):
            value = Decimal(str(value))
        
        # Truncar (não arredondar)
        multiplier = Decimal(10) ** decimals
        truncated = (value * multiplier).to_integral_value(rounding=ROUND_DOWN) / multiplier
        
        return truncated
    
    def calcular_amortizacao(self, valor_principal: float, chave: str) -> Dict:
        """
        Calcula a amortização usando a fórmula: TRUNCAR(G * PROCV(D; 'Amort TS'!G3:H140; 2; 0); 8)
        
        Args:
            valor_principal: Valor G (valor principal)
            chave: Valor D (chave para busca na aba Amort TS)
            
        Returns:
            Dicionário com os resultados do cálculo
        """
        resultado = {
            'sucesso': False,
            'valor_principal': valor_principal,
            'chave': chave,
            'porcentagem': None,
            'resultado': None,
            'erro': None
        }
        
        try:
            # Validar entrada
            if valor_principal is None:
                raise ValueError("valor_principal não pode ser None")
            
            if chave is None or chave == '':
                raise ValueError("chave não pode ser vazia")
            
            # Converter valor_principal para Decimal
            valor_principal_decimal = Decimal(str(valor_principal))
            
            # Buscar porcentagem na aba Amort TS (PROCV)
            porcentagem = self.procv_amort_ts(chave)
            
            if porcentagem is None:
                raise ValueError(f"Chave '{chave}' não encontrada na aba 'Amort TS'")
            
            resultado['porcentagem'] = porcentagem
            
            # Converter porcentagem para Decimal
            porcentagem_decimal = Decimal(str(porcentagem))
            
            # Calcular: valor_principal * porcentagem
            calculo_bruto = valor_principal_decimal * porcentagem_decimal
            
            # Truncar para 8 casas decimais
            resultado_truncado = self.truncate_decimal(calculo_bruto, 8)
            
            resultado['resultado'] = str(resultado_truncado)
            resultado['sucesso'] = True
            
            print(f"✅ Cálculo de amortização realizado com sucesso")
            print(f"   Valor Principal: {valor_principal}")
            print(f"   Chave: {chave}")
            print(f"   Porcentagem: {porcentagem}")
            print(f"   Resultado: {resultado_truncado}")
            
        except ValueError as e:
            resultado['erro'] = str(e)
            print(f"❌ Erro de validação: {resultado['erro']}")
        except Exception as e:
            resultado['erro'] = f"Erro ao calcular amortização: {str(e)}"
            print(f"❌ {resultado['erro']}")
        
        return resultado
    
    def close(self):
        """Fecha a planilha"""
        if self.wb:
            self.wb.close()


def criar_calculadora_amortizacao(planilha_path: str) -> AmortizacaoCalculator:
    """
    Factory function para criar uma instância da calculadora de amortização
    
    Args:
        planilha_path: Caminho para a planilha
        
    Returns:
        Instância de AmortizacaoCalculator
    """
    return AmortizacaoCalculator(planilha_path)