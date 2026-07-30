"""
Calculadora de Juros - Versão 2
Busca dados na planilha CaracterísticasdeCRI-LOGOS-v4-Operacoes.xlsx
Código: Coluna E
Emissão: Coluna X
Taxa: Coluna AF
"""

from decimal import Decimal, ROUND_DOWN
from openpyxl import load_workbook
from datetime import datetime
import os


class JurosCalculator:
    """Calcula juros usando dados da planilha de características"""
    
    def __init__(self, planilha_path=None):
        """
        Inicializa o calculador de juros
        
        Args:
            planilha_path: Caminho da planilha CaracterísticasdeCRI-LOGOS-v4-Operacoes.xlsx
        """
        if planilha_path is None:
            # Procura a planilha na pasta do projeto
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            # Procurar pelo arquivo com nome contendo 'Caracteristicas'
            for arquivo in os.listdir(base_dir):
                if 'Caracteristicas' in arquivo or 'Característica' in arquivo:
                    planilha_path = os.path.join(base_dir, arquivo)
                    break
        
        self.planilha_path = planilha_path
        self.aba_nome = None  # Será detectada automaticamente
    
    def buscar_dados_ativo(self, codigo_if):
        """
        Busca os dados do ativo na planilha
        
        Args:
            codigo_if: Código IF do ativo (ex: 14J0107228)
            
        Returns:
            dict com 'codigo', 'emissao', 'taxa' ou None se não encontrar
        """
        try:
            wb = load_workbook(self.planilha_path, data_only=True)
            # Usar a primeira aba se não foi definida
            aba_nome = self.aba_nome if self.aba_nome else wb.sheetnames[0]
            ws = wb[aba_nome]
            
            # Procurar o código na coluna E
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
                if row[0].value == codigo_if:
                    row_idx = row[0].row
                    
                    # Extrair dados
                    codigo = ws[f'E{row_idx}'].value
                    emissao = ws[f'X{row_idx}'].value
                    taxa = ws[f'AF{row_idx}'].value
                    
                    # Converter emissão para float se necessário
                    if isinstance(emissao, str):
                        # Se for fórmula como "=W3*U3", pular
                        if emissao.startswith('='):
                            return None
                        try:
                            emissao = float(emissao)
                        except:
                            return None
                    
                    # Converter taxa para float
                    if taxa is not None:
                        taxa = float(taxa)
                    
                    wb.close()
                    
                    return {
                        'codigo': codigo,
                        'emissao': emissao,
                        'taxa': taxa
                    }
            
            wb.close()
            return None
            
        except Exception as e:
            print(f"Erro ao buscar dados do ativo: {e}")
            return None
    
    def calcular_juros(self, codigo_if, data_referencia=None):
        """
        Calcula os juros para um ativo
        
        Args:
            codigo_if: Código IF do ativo
            data_referencia: Data de referência (padrão: hoje)
            
        Returns:
            dict com resultado do cálculo ou erro
        """
        # Buscar dados do ativo
        dados = self.buscar_dados_ativo(codigo_if)
        
        if dados is None:
            return {
                'success': False,
                'error': f'Ativo {codigo_if} não encontrado'
            }
        
        try:
            # Extrair valores
            emissao = Decimal(str(dados['emissao']))
            taxa = Decimal(str(dados['taxa']))
            
            # Converter taxa de percentual para decimal (ex: 9.0 -> 0.09)
            taxa_decimal = taxa / Decimal('100')
            
            # Calcular juros: emissão * taxa
            juros = emissao * taxa_decimal
            
            # Truncar para 8 casas decimais
            juros_truncado = juros.quantize(
                Decimal('0.00000001'),
                rounding=ROUND_DOWN
            )
            
            # Data de cálculo
            if data_referencia is None:
                data_calculo = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            else:
                data_calculo = data_referencia
            
            return {
                'success': True,
                'codigo': dados['codigo'],
                'emissao': float(emissao),
                'taxa': float(taxa),
                'juros': float(juros_truncado),
                'data_calculo': data_calculo
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'Erro ao calcular juros: {str(e)}'
            }
    
    def calcular_todos_ativos(self):
        """
        Calcula juros para todos os ativos da planilha
        
        Returns:
            list de dicts com resultados
        """
        try:
            wb = load_workbook(self.planilha_path, data_only=True)
            # Usar a primeira aba se não foi definida
            aba_nome = self.aba_nome if self.aba_nome else wb.sheetnames[0]
            ws = wb[aba_nome]
            
            resultados = []
            data_calculo = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            
            # Iterar sobre todos os ativos
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
                codigo = row[0].value
                
                if codigo:
                    resultado = self.calcular_juros(codigo, data_calculo)
                    if resultado['success']:
                        resultados.append(resultado)
            
            wb.close()
            return resultados
            
        except Exception as e:
            print(f"Erro ao calcular todos os ativos: {e}")
            return []


# Teste rápido
if __name__ == '__main__':
    calc = JurosCalculator()
    
    # Teste 1: Buscar um ativo específico
    print("=== Teste 1: Buscar ativo 14J0107228 ===")
    resultado = calc.calcular_juros('14J0107228')
    print(resultado)
    
    # Teste 2: Calcular todos os ativos
    print("\n=== Teste 2: Calcular todos os ativos ===")
    todos = calc.calcular_todos_ativos()
    for r in todos[:3]:  # Mostrar apenas os 3 primeiros
        print(r)