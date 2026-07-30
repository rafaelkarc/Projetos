import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import os


class AssetIdentifier:
    """
    Classe responsável por interagir com planilhas Excel para identificar e buscar ativos.
    """

    def _load_workbook(self, file_path: str):
        """Carrega a pasta de trabalho (workbook) e trata o erro de arquivo não encontrado."""
        if not os.path.exists(file_path):
            raise FileNotFoundError(
                f"A planilha '{file_path}' não foi encontrada no servidor.")
        return openpyxl.load_workbook(file_path, data_only=True)

    def _get_sheet(self, workbook: openpyxl.workbook.workbook.Workbook, sheet_name: str):
        """Obtém a aba (sheet) e trata o erro de aba não encontrada."""
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"A aba '{sheet_name}' não foi encontrada na planilha. Abas disponíveis: {workbook.sheetnames}")
        return workbook[sheet_name]

    def read_cell_value(self, file_path: str, sheet_name: str, cell_address: str):
        """Lê e retorna o valor de uma célula específica em uma planilha.

        Args:
            file_path (str): O caminho para o arquivo Excel.
            sheet_name (str): O nome da aba na planilha.
            cell_address (str): O endereço da célula (ex: 'A1').

        Returns:
            dict: Um dicionário contendo o status da operação, o valor da célula e metadados.
        """
        try:
            workbook = self._load_workbook(file_path)
            sheet = self._get_sheet(workbook, sheet_name)
            cell = sheet[cell_address]
            cell_value = cell.value
            value_type = type(cell_value).__name__
            if isinstance(cell_value, datetime):
                cell_value = cell_value.isoformat()
            return {
                "success": True, "value": cell_value, "value_type": value_type,
                "file_path": file_path, "sheet_name": sheet_name, "cell_address": cell_address,
                "timestamp": datetime.now().isoformat()
            }
        except (FileNotFoundError, ValueError, KeyError) as e:
            return {"success": False, "error": str(e)}
        except Exception as e:
            return {"success": False, "error": f"Ocorreu um erro inesperado ao ler a célula: {str(e)}"}

    def get_sheet_info(self, file_path: str):
        """Lista todas as abas (sheets) em uma planilha e retorna suas informações.

        Args:
            file_path (str): O caminho para o arquivo Excel.

        Returns:
            dict: Um dicionário contendo o status da operação e informações sobre as abas.
        """
        try:
            workbook = self._load_workbook(file_path)
            sheets_info = [{"name": sheet.title, "max_row": sheet.max_row,
                            "max_column": sheet.max_column} for sheet in workbook.worksheets]
            return {
                "success": True, "file_path": file_path, "total_sheets": len(sheets_info),
                "sheets": sheets_info, "timestamp": datetime.now().isoformat()
            }
        except (FileNotFoundError, ValueError) as e:
            return {"success": False, "error": str(e)}
        except Exception as e:
            return {"success": False, "error": f"Ocorreu um erro inesperado ao listar as abas: {str(e)}"}

    def search_value_in_range(self, file_path: str, sheet_name: str, start_cell: str, end_cell: str, search_value: any):
        """Busca um valor específico dentro de um intervalo de células em uma planilha.

        Args:
            file_path (str): O caminho para o arquivo Excel.
            sheet_name (str): O nome da aba na planilha.
            start_cell (str): A célula inicial do intervalo (ex: 'A1').
            end_cell (str): A célula final do intervalo (ex: 'C10').
            search_value (any): O valor a ser buscado.

        Returns:
            dict: Um dicionário contendo o status da operação, o número de correspondências e os detalhes das correspondências.
        """
        try:
            workbook = self._load_workbook(file_path)
            sheet = self._get_sheet(workbook, sheet_name)
            search_range = sheet[f'{start_cell}:{end_cell}']
            matches = []
            for row in search_range:
                for cell in row:
                    if cell.value is not None and str(cell.value).strip().upper() == str(search_value).strip().upper():
                        matches.append({"cell_address": cell.coordinate, "value": cell.value,
                                       "row": cell.row, "column_letter": get_column_letter(cell.column)})
            return {
                "success": True, "search_value": search_value, "total_matches": len(matches),
                "matches": matches, "timestamp": datetime.now().isoformat(),
                "message": "Busca concluída." if matches else "Busca concluída, mas o valor não foi encontrado."
            }
        except (FileNotFoundError, ValueError) as e:
            return {"success": False, "error": str(e)}
        except Exception as e:
            return {"success": False, "error": f"Ocorreu um erro inesperado ao buscar na planilha: {str(e)}"}

    def read_column_values(self, file_path: str, sheet_name: str, column: str, start_row: int = 1, end_row: int = None):
        """Lê e retorna todos os valores de uma coluna específica.

        Args:
            file_path (str): O caminho para o arquivo Excel.
            sheet_name (str): O nome da aba na planilha.
            column (str): A letra da coluna (ex: 'A', 'B').
            start_row (int, optional): A linha inicial para leitura. Padrão é 1.
            end_row (int, optional): A linha final para leitura. Padrão é a última linha da coluna.

        Returns:
            dict: Um dicionário contendo o status da operação e os valores da coluna.
        """
        try:
            workbook = self._load_workbook(file_path)
            sheet = self._get_sheet(workbook, sheet_name)

            # Converte a letra da coluna para um índice numérico
            column_index = column_index_from_string(column.upper())

            # Define a linha final se não for especificada
            if end_row is None:
                end_row = sheet.max_row

            values = []
            # Itera sobre as linhas da coluna especificada
            for row_num in range(start_row, end_row + 1):
                cell = sheet.cell(row=row_num, column=column_index)
                cell_value = cell.value
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.isoformat()

                values.append({
                    "row": row_num,
                    "cell_address": f"{column.upper()}{row_num}",
                    "value": cell_value
                })

            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "column": column.upper(),
                "start_row": start_row,
                "end_row": end_row,
                "total_values": len(values),
                "values": values,
                "timestamp": datetime.now().isoformat()
            }
        except (FileNotFoundError, ValueError) as e:
            return {"success": False, "error": str(e)}
        except Exception as e:
            return {"success": False, "error": f"Ocorreu um erro inesperado ao ler a coluna: {str(e)}"}

    def buscar_por_codigo_if(self, file_path: str, search_value: any):
        """
        Busca um ativo na coluna E da aba principal da planilha pelo código IF,
        retornando também a descrição da coluna (linha 1 da planilha).

        Args:
            file_path (str): O caminho para o arquivo Excel.
            search_value (any): O código IF a ser buscado.

        Returns:
            dict: Um dicionário contendo o status da operação e os resultados da busca.
        """
        try:
            workbook = self._load_workbook(file_path)
            sheet = workbook[workbook.sheetnames[0]]  # Assume a primeira aba

            # Captura o cabeçalho da planilha (linha 1)
            headers = [cell.value for cell in sheet[1]]

            # Define índice da coluna E (5ª coluna, índice 4)
            target_col_index = 4

            # Garante que a descrição da coluna existe
            descricao_coluna = (
                headers[target_col_index]
                if len(headers) > target_col_index and headers[target_col_index]
                else "Coluna E"
            )

            resultados = []
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):  # pula o cabeçalho
                cell = row[target_col_index]
                if cell.value and str(cell.value).strip().upper() == str(search_value).strip().upper():
                    linha = [c.value for c in row]
                    resultados.append({
                        "linha": cell.row,
                        "coluna": "E",
                        "coluna_descricao": descricao_coluna,
                        "valores": linha
                    })

            if resultados:
                return {"success": True, "matches": resultados}
            else:
                return {"success": True, "matches": [], "message": "Código IF não encontrado."}

        except (FileNotFoundError, ValueError) as e:
            return {"success": False, "error": str(e)}
        except Exception as e:
            return {"success": False, "error": f"Erro ao buscar na planilha: {str(e)}"}


# Criar uma instância para ser importada em outros módulos
asset_identifier = AssetIdentifier()